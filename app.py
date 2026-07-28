from flask import jsonify
from flask import Flask, render_template, request, redirect, session, send_file, url_for, flash
from functools import wraps
import pandas as pd
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
import os
from io import BytesIO, StringIO
import json
from datetime import datetime, timedelta
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from werkzeug.utils import secure_filename
import re
from PIL import Image
import cv2
import pdfplumber
import openpyxl
import google.generativeai as genai

app = Flask(__name__)
app.secret_key = "super_secret_key"
BASE_DIR = os.path.abspath(os.path.dirname(__file__)) 
DB_PATH = os.path.join(BASE_DIR, "data", "users.db")
os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {"pdf", "csv", "xls", "xlsx"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
GEMINI_API_KEY = "AIzaSyBuEuYpLw3cUG09Ejzr9iSbgAzIYWqJpXk"

EXPENSE_CATEGORIES = [
    {"name": "Home & Living", "icon": "home", "color": "#8b5cf6", "subcategories": ["Rent", "House Maintenance", "Furniture", "Interior/Decor", "Cleaning Supplies", "Maid Salary", "Gas Cylinder", "Water Bill", "Electricity Bill", "Internet/Wi-Fi", "Mobile Recharge", "DTH/Cable TV"]},
    {"name": "Food", "icon": "food", "color": "#f97316", "subcategories": ["Groceries", "Vegetables & Fruits", "Dairy Products", "Snacks", "Dining Out", "Coffee/Tea", "Food Delivery", "Bakery"]},
    {"name": "Transport", "icon": "car", "color": "#2563eb", "subcategories": ["Fuel/Petrol", "Bus", "Train", "Metro", "Auto/Taxi", "Cab Apps", "Parking Fees", "Toll Charges", "Vehicle Service", "Vehicle Insurance", "Vehicle EMI"]},
    {"name": "Shopping", "icon": "bag", "color": "#db2777", "subcategories": ["Clothes", "Footwear", "Accessories", "Cosmetics", "Electronics", "Gadgets", "Home Appliances"]},
    {"name": "Health", "icon": "health", "color": "#16a34a", "subcategories": ["Medicines", "Doctor Consultation", "Hospital Bills", "Lab Tests", "Health Insurance", "Gym Membership", "Skincare", "Mental Wellness"]},
    {"name": "Education", "icon": "book", "color": "#0891b2", "subcategories": ["Tuition Fees", "Course Purchases", "Books", "Certifications", "Exam Fees", "College Expenses", "Stationery"]},
    {"name": "Entertainment", "icon": "play", "color": "#7c3aed", "subcategories": ["Movies", "OTT Subscriptions", "Games", "Music Apps", "Trips", "Parties", "Hobbies"]},
    {"name": "Financial", "icon": "bank", "color": "#0f766e", "subcategories": ["Loan EMI", "Credit Card Bill", "Tax", "Investments", "SIP", "Mutual Funds", "Insurance Premium", "Savings Transfer"]},
    {"name": "Family & Personal", "icon": "gift", "color": "#ec4899", "subcategories": ["Gifts", "Donations", "Family Support", "Child Expenses", "Pet Expenses", "Personal Care", "Salon/Beauty"]},
    {"name": "Work & Business", "icon": "work", "color": "#475569", "subcategories": ["Office Supplies", "Software Subscriptions", "Cloud Services", "Domain & Hosting", "Freelancers", "Marketing", "Travel for Work"]},
    {"name": "Subscription Tracking", "icon": "repeat", "color": "#9333ea", "subcategories": ["Netflix", "Spotify", "YouTube Premium", "ChatGPT", "Adobe", "Canva"]},
    {"name": "Event Expenses", "icon": "calendar", "color": "#e11d48", "subcategories": ["Wedding", "Birthday", "Festival Shopping", "Vacation", "College Events"]},
    {"name": "Emergency Expenses", "icon": "alert", "color": "#dc2626", "subcategories": ["Emergency Medical", "Device Repair", "Accident Repair", "Unexpected Costs"]},
]

for category in EXPENSE_CATEGORIES:
    category["slug"] = re.sub(r"[^a-z0-9]+", "-", category["name"].lower()).strip("-")

INCOME_CATEGORIES = ["Salary", "Business", "Rental", "Investment", "Gifts", "Others"]


def find_expense_category(slug):
    return next((category for category in EXPENSE_CATEGORIES if category["slug"] == slug), None)


def parse_transaction_date(raw_timestamp):
    try:
        return datetime.fromisoformat(str(raw_timestamp or "").replace("Z", "")).date()
    except ValueError:
        return None


def build_insights(username: str):
    today = datetime.today().date()
    month_start = today.replace(day=1)
    previous_month_end = month_start - timedelta(days=1)
    previous_month_start = previous_month_end.replace(day=1)

    with get_conn(row_factory=sqlite3.Row) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT description, amount, COALESCE(LOWER(type),'expense') AS type,
                   COALESCE(category,'Uncategorized') AS category,
                   COALESCE(subcategory,'Other') AS subcategory,
                   timestamp
            FROM expenses
            WHERE user = ?
        """, (username,))
        rows = cursor.fetchall()

    current_income = current_expense = current_saving = 0.0
    previous_income = previous_expense = previous_saving = 0.0
    current_expense_by_category = {}
    previous_expense_by_category = {}
    current_income_by_category = {}

    for row in rows:
        tx_date = parse_transaction_date(row["timestamp"])
        if not tx_date:
            continue

        amount = float(row["amount"] or 0)
        tx_type = (row["type"] or "expense").lower()
        category = row["category"] or row["subcategory"] or "Other"

        if month_start <= tx_date <= today:
            if tx_type == "income":
                current_income += amount
                current_income_by_category[category] = current_income_by_category.get(category, 0.0) + amount
            elif tx_type == "saving":
                current_saving += amount
            else:
                current_expense += amount
                current_expense_by_category[category] = current_expense_by_category.get(category, 0.0) + amount
        elif previous_month_start <= tx_date <= previous_month_end:
            if tx_type == "income":
                previous_income += amount
            elif tx_type == "saving":
                previous_saving += amount
            else:
                previous_expense += amount
                previous_expense_by_category[category] = previous_expense_by_category.get(category, 0.0) + amount

    current_saved = max(0, current_income - current_expense)
    previous_saved = max(0, previous_income - previous_expense)
    spent_percent = round((current_expense / current_income) * 100) if current_income > 0 else 0
    savings_rate = round((current_saved / current_income) * 100) if current_income > 0 else 0
    income_growth = round(((current_income - previous_income) / previous_income) * 100) if previous_income > 0 else 0
    net_worth_growth = round(current_saved - previous_saved, 2)

    top_expense = max(current_expense_by_category.items(), key=lambda item: item[1], default=("Spending", 0))
    previous_top = previous_expense_by_category.get(top_expense[0], 0)
    if previous_top > 0:
        expense_change = round(((top_expense[1] - previous_top) / previous_top) * 100)
        top_expense_line = f"{top_expense[0]} spending {'increased' if expense_change >= 0 else 'decreased'} by {abs(expense_change)}%"
    elif top_expense[1] > 0:
        top_expense_line = f"{top_expense[0]} is your top spending category this month"
    else:
        top_expense_line = "Add expenses to see spending behaviour"

    category_lines = []
    for name in ["Food", "Transport", "Shopping"]:
        current_value = current_expense_by_category.get(name, 0)
        previous_value = previous_expense_by_category.get(name, 0)
        if previous_value > 0:
            change = round(((current_value - previous_value) / previous_value) * 100)
            category_lines.append(f"{name} spending {'increased' if change >= 0 else 'decreased'} by {abs(change)}%")
        elif current_value > 0:
            category_lines.append(f"{name} spending is Rs. {round(current_value, 2)} this month")
        else:
            category_lines.append(f"{name} has no spending this month")

    income_lines = []
    for category, amount in sorted(current_income_by_category.items(), key=lambda item: item[1], reverse=True)[:4]:
        percent = round((amount / current_income) * 100) if current_income > 0 else 0
        income_lines.append(f"{category} contributes {percent}% of income")
    if not income_lines:
        income_lines.append("Add income to see income analysis")

    home_insight = "Add income and expenses to unlock this month's insight."
    if current_income > 0:
        home_insight = f"You saved {savings_rate}% of your income this month."

    return {
        "home": home_insight,
        "sections": [
            {"title": "This Month", "items": [
                f"You saved Rs. {round(current_saved, 2)}",
                f"You spent {spent_percent}% of your income",
                "You are saving more than last month" if current_saved >= previous_saved else "You saved less than last month"
            ]},
            {"title": "Spending Behaviour", "items": [top_expense_line] + category_lines},
            {"title": "Income Analysis", "items": income_lines},
            {"title": "Financial Growth", "items": [
                f"Current savings rate: {savings_rate}%",
                f"Income {'increased' if income_growth >= 0 else 'decreased'} by {abs(income_growth)}%",
                f"Net worth {'increased' if net_worth_growth >= 0 else 'decreased'} by Rs. {abs(net_worth_growth)}"
            ]}
        ]
    }


def build_goal_rows(username: str):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, goal_name, target_amount, saved_amount, target_date
            FROM goals
            WHERE user = ?
            ORDER BY id DESC
        """, (username,))
        rows = cursor.fetchall()

    goals = []
    today = datetime.today().date()

    for row in rows:
        goal_id, goal_name, target_amount, saved_amount, target_date = row
        target_amount = float(target_amount or 0)
        saved_amount = float(saved_amount or 0)

        try:
            target_date_obj = datetime.strptime(target_date, "%Y-%m-%d").date()
            days_left = (target_date_obj - today).days
        except (TypeError, ValueError):
            days_left = 0

        remaining_amount = max(0, target_amount - saved_amount)
        if days_left > 0:
            daily_required = remaining_amount / days_left
            monthly_required = (remaining_amount / days_left) * 30
        else:
            daily_required = remaining_amount
            monthly_required = remaining_amount

        progress = 0
        if target_amount > 0:
            progress = min(100, round((saved_amount / target_amount) * 100, 2))

        goals.append({
            "id": goal_id,
            "goal_name": goal_name,
            "target_amount": round(target_amount, 2),
            "saved_amount": round(saved_amount, 2),
            "target_date": target_date,
            "days_left": days_left,
            "remaining_amount": round(remaining_amount, 2),
            "daily_required": round(daily_required, 2),
            "monthly_required": round(monthly_required, 2),
            "progress": progress
        })

    return goals


def get_conn(row_factory=None):
    conn = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
    if row_factory:
        conn.row_factory = row_factory
    return conn
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def fetch_user_totals(username: str):
    """
    Returns totals for the logged-in user:
    savings are money set aside, so they reduce available balance.
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(LOWER(type),'expense') AS ttype, COALESCE(SUM(amount),0)
            FROM expenses
            WHERE user = ?
            GROUP BY COALESCE(LOWER(type),'expense')
        """, (username,))
        rows = cur.fetchall()

    totals = {"income": 0.0, "saving": 0.0, "expense": 0.0}
    for ttype, s in rows:
        ttype = (ttype or "expense").lower()
        s = float(s or 0)
        if ttype == "income":
            totals["income"] += s
        elif ttype == "saving":
            totals["saving"] += s
        else:
            totals["expense"] += s

    # ✅ requirement: saving should reflect in total income
    income_total = totals["income"]
    balance = income_total - totals["expense"] - totals["saving"]

    return {
        "income_total": round(income_total, 2),
        "income_only": round(totals["income"], 2),
        "saving": round(totals["saving"], 2),
        "expense": round(totals["expense"], 2),
        "balance": round(balance, 2)
    }


def fetch_available_balance(username: str) -> float:
    return float(fetch_user_totals(username)["balance"])


def fetch_recent_transactions(username: str, limit: int = 5):
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT description, amount, COALESCE(LOWER(type),'expense') AS ttype,
                   COALESCE(category,'Uncategorized') AS category,
                   timestamp
            FROM expenses
            WHERE user = ?
            ORDER BY timestamp DESC
            LIMIT ?
        """, (username, limit))
        rows = cur.fetchall()

    tx = []
    for desc, amt, ttype, cat, ts in rows:
        tx.append({
            "description": desc or "",
            "amount": float(amt or 0),
            "type": (ttype or "expense"),
            "category": cat or "Uncategorized",
            "timestamp": ts
        })
    return tx


def fetch_category_summary(username: str, for_type: str = "expense"):
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(category,'Uncategorized') AS category, COALESCE(SUM(amount),0)
            FROM expenses
            WHERE user = ? AND COALESCE(LOWER(type),'expense') = ?
            GROUP BY COALESCE(category,'Uncategorized')
            ORDER BY 2 DESC
        """, (username, for_type.lower()))
        rows = cur.fetchall()

    return [{"category": c, "amount": round(float(a or 0), 2)} for c, a in rows]

# Ensure expenses table exists
with get_conn() as conn:
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user TEXT NOT NULL,
            description TEXT,
            amount REAL,
            category TEXT DEFAULT 'Uncategorized',
            emotion TEXT,
            type TEXT DEFAULT 'expense',
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute("PRAGMA table_info(expenses)")
    existing_columns = {row[1] for row in cursor.fetchall()}
    if "subcategory" not in existing_columns:
        cursor.execute("ALTER TABLE expenses ADD COLUMN subcategory TEXT DEFAULT ''")
    conn.commit()

# Ensure users table exists
with get_conn() as conn:
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            contact TEXT
        )
    ''')
    conn.commit()

    # Ensure goals table exists
with get_conn() as conn:
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS goals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user TEXT NOT NULL,
            goal_name TEXT NOT NULL,
            target_amount REAL NOT NULL,
            saved_amount REAL DEFAULT 0,
            target_date TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()


# ------------------------- Helpers -------------------------
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:   # if no user session, redirect
            flash("Please login first!", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


# ------------------------- Routes -------------------------


#-------------------
@app.route("/analysis")
@login_required
def analysis():
    user = session["user"]

    # Example values (replace with your DB aggregations)
    totals = {"income": 8700, "expense": 992.95, "balance": 7707.05}

    week = [
        {"day":"Sun","expense":"-416.2"},
        {"day":"Mon","expense":"-218.7"},
        {"day":"Tue","expense":"-175.5"},
        {"day":"Wed","expense":"-17.2"},
        {"day":"Thu","expense":"-95.0"},
        {"day":"Fri","expense":"-16.2"},
        {"day":"Sat","expense":"-54.3"},
    ]

    grouped = [
        {"date":"Jan 03, Sunday", "items":[
            {"icon":"👕","category":"Clothing","account":"Card","amount":"65.55","type":"expense"},
            {"icon":"🛒","category":"Shopping","account":"Card","amount":"120.00","type":"expense"},
        ]},
        {"date":"Jan 02, Saturday", "items":[
            {"icon":"🍿","category":"Entertainment","account":"Card","amount":"30.15","type":"expense"},
        ]},
    ]

    flow_points = json.dumps([420, 260, 210, 95, 120, 60, 80])

    return render_template(
        "analysis.html",
        username=user,
        totals=totals,
        week=week,
        grouped=grouped,
        flow_points=flow_points
    )

#----------------------------------
@app.route("/budgets")
@login_required
def budgets(): return "Budgets page (coming soon)"

@app.route("/accounts")
@login_required
def accounts(): return "Accounts page (coming soon)"

@app.route("/categories")
@login_required
def categories(): return "Categories page (coming soon)"
#-------------------------------------------------------------
@app.route('/')
def home():
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        contact = request.form.get('contact', '').strip()  # Optional email/contact

        hashed_password = generate_password_hash(password)

        try:
            with sqlite3.connect(DB_PATH) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO users (username, password, contact) VALUES (?, ?, ?)",
                    (username, hashed_password, contact)
                )
                conn.commit()
                flash("✅ Registration successful! Please login.", "success")
                return redirect('/login')
        except sqlite3.IntegrityError:
            flash("❌ Username already exists!", "danger")

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']

        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
            user = cursor.fetchone()

        if user and check_password_hash(user[2], password):  # user[2] = password
            session['user'] = username
            flash("✅ Login successful!", "success")
            return redirect(url_for('dashboard'))
        else:
            flash("❌ Invalid username or password", "danger")
            return redirect('/login')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('user', None)
    flash("Logged out successfully.", "info")
    return redirect('/login')


@app.route("/profile")
@login_required
def profile():
    return render_template("profile.html", user=session["user"])
# ---------------- Dashboard ----------------
@app.route('/dashboard')
@login_required
def dashboard():
    user = session['user']

    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, description, amount, emotion, timestamp, type,
                   COALESCE(category,'Uncategorized') AS category,
                   COALESCE(subcategory,'') AS subcategory
            FROM expenses
            WHERE user = ?
            ORDER BY timestamp DESC
        """, (user,))
        rows = cursor.fetchall()

    expenses = []
    total_spent = 0.0
    total_income = 0.0
    total_savings = 0.0

    for row in rows:
        entry_id, desc, amt, emo, time, ttype, category, subcategory = row

        try:
            amt_val = float(amt) if amt is not None else 0.0
        except (ValueError, TypeError):
            amt_val = 0.0

        ttype = (ttype or "").lower()

        expenses.append({
            "id": entry_id,
            "description": desc,
            "amount": amt_val,
            "emotion": emo,
            "timestamp": time,
            "type": ttype,
            "category": category,
            "subcategory": subcategory
        })

        if ttype == "income":
            total_income += amt_val
        elif ttype == "saving":
            total_savings += amt_val
        elif ttype == "expense":
            total_spent += amt_val

    # ✅ Balance
    balance = total_income - total_spent - total_savings

    # ✅ Improved Wealth Score
    if total_income > 0:
        savings_ratio = total_savings / total_income
        expense_ratio = total_spent / total_income
        balance_ratio = max(0, balance / total_income)

        wealth_score = (
            savings_ratio * 0.35 +
            (1 - min(expense_ratio, 1)) * 0.35 +
            balance_ratio * 0.30
        ) * 100

        wealth_score = max(0, min(100, round(wealth_score)))
    else:
        wealth_score = 0

    # ✅ Wealth Status
    if wealth_score >= 80:
        wealth_status = "🟢 Excellent"
    elif wealth_score >= 60:
        wealth_status = "🟡 Good"
    elif wealth_score >= 40:
        wealth_status = "🟠 Average"
    else:
        wealth_status = "🔴 Poor"

    return render_template(
        "dashboard.html",
        username=user,
        expenses=expenses,
        total=round(total_spent, 2),
        income=round(total_income, 2),
        savings=round(total_savings, 2),
        balance=round(balance, 2),
        health_score=wealth_score,
        health_status=wealth_status,
        score=wealth_score,
        behavior=wealth_status,
        expense_categories=EXPENSE_CATEGORIES,
        insight=build_insights(user)["home"]
    )


@app.route('/insights')
@login_required
def insights():
    user = session['user']
    return render_template("insights.html", username=user, insights=build_insights(user))
# ---------------- Expenses Page ----------------
@app.route('/expenses')
@login_required
def expenses():
    user = session.get("user")

    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT description, amount, type, timestamp FROM expenses WHERE user = ?", (user,))
        rows = cursor.fetchall()

    entries = []
    for r in rows:
        entries.append({
            "description": r[0],
            "amount": r[1],
            "type": r[2],
            "date": r[3]
        })

    return render_template("expenses.html", entries=entries)


# ---------------- Visualization ----------------
@app.route('/visualization')
@login_required
def visualization():
    user = session['user']

    with get_conn() as conn:
        cursor = conn.cursor()

        # totals for table + donut
        cursor.execute("""
            SELECT LOWER(type), SUM(amount)
            FROM expenses
            WHERE user = ?
            GROUP BY LOWER(type)
        """, (user,))
        rows = cursor.fetchall()

        total_income = 0.0
        total_expense = 0.0
        total_saving = 0.0

        for ttype, total in rows:
            total = float(total or 0)
            if ttype == "income":
                total_income += total
            elif ttype == "saving":
                total_saving += total
            elif ttype == "expense":
                total_expense += total

        balance = total_income - total_expense - total_saving

        # month-wise graph data
        cursor.execute("""
            SELECT strftime('%m', timestamp) as month, LOWER(type), SUM(amount)
            FROM expenses
            WHERE user = ?
            GROUP BY strftime('%m', timestamp), LOWER(type)
            ORDER BY month
        """, (user,))
        monthly_rows = cursor.fetchall()

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    income_data = [0] * 12
    expense_data = [0] * 12

    for month, ttype, total in monthly_rows:
        idx = int(month) - 1
        total = float(total or 0)

        if ttype == "income":
            income_data[idx] += total
        elif ttype == "expense":
            expense_data[idx] += total

    return render_template(
        "visualization.html",
        username=user,
        income=round(total_income, 2),
        expense=round(total_expense, 2),
        saving=round(total_saving, 2),
        balance=round(balance, 2),
        month_labels=months,
        income_data=income_data,
        expense_data=expense_data
    )
# ---------------- Calendar ----------------
@app.route('/calendar')
@login_required
def calendar_view():
    user = session['user']

    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT description, amount, emotion, timestamp, type, category
            FROM expenses
            WHERE user = ?
            ORDER BY timestamp ASC
        """, (user,))
        rows = cursor.fetchall()

    events = []
    income = 0.0
    expense = 0.0
    saving = 0.0

    for desc, amt, emo, date, ttype, category in rows:
        try:
            amt_val = float(amt) if amt is not None else 0.0
        except (ValueError, TypeError):
            amt_val = 0.0

        ttype = (ttype or "").lower()

        if ttype == "income":
            color = "#047857"
            income += amt_val
        elif ttype == "saving":
            color = "#0ea5e9"
            saving += amt_val
        else:
            color = "#b45309"
            expense += amt_val

        # ✅ make title exactly like food:200
        if amt_val.is_integer():
            amount_text = str(int(amt_val))
        else:
            amount_text = str(amt_val)

        type_label = ttype.title() if ttype else "Expense"
        title_text = f"{type_label} Rs. {amount_text}"

        # ✅ only date part should go to FullCalendar
        only_date = str(date)[:10]

        events.append({
            "title": title_text,
            "start": only_date,
            "allDay": True,
            "color": color,
            "classNames": [f"cal-{ttype or 'expense'}"],
            "extendedProps": {
                "type": type_label,
                "category": category or "Uncategorized",
                "description": desc or type_label,
                "amount": amount_text,
                "date": only_date
            }
        })

    balance = income - expense - saving

    return render_template(
        "calendar.html",
        username=user,
        events=events,   # ✅ pass Python list directly
        income=round(income, 2),
        expenses=round(expense, 2),
        savings=round(saving, 2),
        balance=round(balance, 2)
    )

#-------------------add entry--------------------
@app.route("/add-transaction")
@login_required
def add_transaction_categories():
    return render_template(
        "add_transaction_categories.html",
        username=session["user"],
        expense_categories=EXPENSE_CATEGORIES,
        income_categories=INCOME_CATEGORIES,
        goals=build_goal_rows(session["user"])
    )


@app.route("/add-transaction/<category_slug>")
@login_required
def add_transaction_subcategories(category_slug):
    category = find_expense_category(category_slug)
    if not category:
        flash("Please choose a valid expense category.", "danger")
        return redirect(url_for("add_transaction_categories"))

    return render_template(
        "add_transaction_subcategories.html",
        username=session["user"],
        category=category
    )


@app.route("/add_entry", methods=["POST"])
@login_required
def add_entry():
    description = (request.form.get("description") or "").strip()
    category = (request.form.get("category") or "").strip()
    subcategory = (request.form.get("subcategory") or "").strip()
    entry_type = (request.form.get("type") or "expense").strip().lower()
    emotion = (request.form.get("emotion") or "Neutral").strip()
    timestamp_raw = (request.form.get("timestamp") or "").strip()

    try:
        amount = float(request.form.get("amount", 0))
    except (ValueError, TypeError):
        flash("Please enter a valid amount.", "danger")
        return redirect(url_for("dashboard") + "#add-entry")

    if amount <= 0:
        flash("Amount must be greater than zero.", "danger")
        return redirect(url_for("dashboard") + "#add-entry")

    if entry_type not in {"income", "expense", "saving"}:
        flash("Please choose a valid transaction type.", "danger")
        return redirect(url_for("dashboard") + "#add-entry")

    username = session["user"]

    if entry_type in {"expense", "saving"}:
        available = fetch_available_balance(username)
        if amount > available:
            flash(
                f"Cannot add Rs. {amount:,.2f}. Your available balance is Rs. {available:,.2f}.",
                "danger"
            )
            return redirect(url_for("dashboard") + "#add-entry")

    if entry_type == "expense" and (not category or not subcategory):
        flash("Please select an expense category and detail.", "danger")
        return redirect(url_for("dashboard") + "#add-entry")

    if not description:
        description = subcategory or category or entry_type.title()

    timestamp_value = None
    if timestamp_raw:
        try:
            timestamp_value = datetime.fromisoformat(timestamp_raw).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            flash("Please choose a valid date and time.", "danger")
            return redirect(url_for("dashboard") + "#add-entry")

    username = session["user"]  # ✅ REQUIRED (NOT NULL)

    with get_conn() as conn:
        cursor = conn.cursor()
        if timestamp_value:
            cursor.execute(
                """
                INSERT INTO expenses (user, description, amount, category, subcategory, emotion, type, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (username, description, amount, category, subcategory, emotion, entry_type, timestamp_value)
            )
        else:
            cursor.execute(
                """
                INSERT INTO expenses (user, description, amount, category, subcategory, emotion, type)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (username, description, amount, category, subcategory, emotion, entry_type)
            )
        conn.commit()

    flash("Transaction saved successfully.", "success")
    return redirect(url_for("dashboard"))


@app.route("/add-goal-saving", methods=["POST"])
@login_required
def add_goal_saving():
    username = session["user"]
    goal_id_raw = request.form.get("goal_id")
    timestamp_raw = (request.form.get("timestamp") or "").strip()

    try:
        goal_id = int(goal_id_raw)
        amount = float(request.form.get("amount", 0))
    except (ValueError, TypeError):
        flash("Please choose a goal and enter a valid amount.", "danger")
        return redirect(url_for("add_transaction_categories"))

    if amount <= 0:
        flash("Goal saving amount must be greater than zero.", "danger")
        return redirect(url_for("add_transaction_categories"))

    available = fetch_available_balance(username)
    if amount > available:
        flash(
            f"Cannot save Rs. {amount:,.2f}. Your available balance is Rs. {available:,.2f}.",
            "danger"
        )
        return redirect(url_for("add_transaction_categories"))

    timestamp_value = None
    if timestamp_raw:
        try:
            timestamp_value = datetime.fromisoformat(timestamp_raw).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            flash("Please choose a valid date and time.", "danger")
            return redirect(url_for("add_transaction_categories"))

    with get_conn(row_factory=sqlite3.Row) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT goal_name FROM goals WHERE id = ? AND user = ?",
            (goal_id, username)
        )
        goal = cursor.fetchone()
        if not goal:
            flash("Please choose a valid goal.", "danger")
            return redirect(url_for("add_transaction_categories"))

        goal_name = goal["goal_name"]
        cursor.execute(
            "UPDATE goals SET saved_amount = COALESCE(saved_amount, 0) + ? WHERE id = ? AND user = ?",
            (amount, goal_id, username)
        )

        if timestamp_value:
            cursor.execute(
                """
                INSERT INTO expenses (user, description, amount, category, subcategory, emotion, type, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (username, f"Goal saving - {goal_name}", amount, "Goal", goal_name, "Neutral", "saving", timestamp_value)
            )
        else:
            cursor.execute(
                """
                INSERT INTO expenses (user, description, amount, category, subcategory, emotion, type)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (username, f"Goal saving - {goal_name}", amount, "Goal", goal_name, "Neutral", "saving")
            )
        conn.commit()

    flash("Goal saving added successfully.", "success")
    return redirect(url_for("goal_quick"))


# ---------------- Monthly Report ----------------
@app.route('/download-report', methods=['POST'])
@login_required
def download_report():
    user = session['user']
    try:
        months = int(request.form.get('months', 1))
    except ValueError:
        months = 1
    file_format = request.form.get('format', 'csv').lower()

    with get_conn(row_factory=sqlite3.Row) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT DATE(timestamp) AS date,
                   COALESCE(type,'expense') AS type,
                   COALESCE(category,'Uncategorized') AS category,
                   description,
                   amount
            FROM expenses
            WHERE user = ?
              AND DATE(timestamp) >= DATE('now', ?)
            ORDER BY timestamp ASC
            """,
            (user, f'-{months} months')
        )
        rows = cur.fetchall()

    if not rows:
        return "No transactions found for the selected period.", 404

    # convert rows to DataFrame
    df = pd.DataFrame([dict(r) for r in rows])
    # ensure required columns exist
    for c in ['date', 'type', 'category', 'description', 'amount']:
        if c not in df.columns:
            df[c] = None

    df['type'] = df['type'].str.lower().fillna('expense')
    df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0.0)

    df['month'] = pd.to_datetime(df['date']).dt.to_period('M').astype(str)
    summary = (
        df.groupby(['month', 'type'])['amount']
          .sum()
          .unstack(fill_value=0)
    )

    for col in ('income', 'expense', 'saving'):
        if col not in summary.columns:
            summary[col] = 0.0

    summary['balance'] = summary['income'] - summary['expense'] - summary['saving']
    summary = summary.reset_index()

    if file_format == 'csv':
        s_io = StringIO()
        s_io.write("FinGuard Monthly Report\n")
        s_io.write(f"User: {user}\n")
        s_io.write(f"Period: last {months} month(s)\n\n")
        s_io.write("Transactions\n")
        df[['date', 'type', 'category', 'description', 'amount']].to_csv(s_io, index=False)
        s_io.write("\nSummary by Month\n")
        summary.to_csv(s_io, index=False)

        data = s_io.getvalue().encode('utf-8')
        return send_file(
            BytesIO(data),
            as_attachment=True,
            download_name=f"FinGuard_Report_{months}m.csv",
            mimetype='text/csv'
        )

    elif file_format == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("FinGuard Monthly Report", styles['Title']))
        story.append(Paragraph(f"User: {user}", styles['Normal']))
        story.append(Paragraph(f"Period: last {months} month(s)", styles['Normal']))
        story.append(Spacer(1, 12))

        # Transactions
        story.append(Paragraph("Transactions", styles['Heading2']))
        tx_data = [['Date', 'Type', 'Category', 'Description', 'Amount']] + \
                  df[['date','type','category','description','amount']].values.tolist()
        tx_table = Table(tx_data, repeatRows=1)
        tx_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('ALIGN', (-1,1), (-1,-1), 'RIGHT'),
        ]))
        story.append(tx_table)
        story.append(Spacer(1, 16))

        # Summary
        story.append(Paragraph("Summary by Month", styles['Heading2']))
        sum_data = [['Month', 'Income', 'Expense', 'Saving', 'Balance']] + \
                   summary[['month','income','expense','saving','balance']].round(2).values.tolist()
        sum_table = Table(sum_data, repeatRows=1)
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
        ]))
        story.append(sum_table)

        doc.build(story)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"FinGuard_Report_{months}m.pdf",
            mimetype='application/pdf'
        )

    else:
        return "Unsupported format.", 400
#------------------------------------------

@app.route("/ai-chat", methods=["POST"])
@login_required
def ai_chat():
    user = session["user"]
    payload = request.get_json(silent=True) or {}
    msg = (payload.get("message") or "").strip()

    if not msg:
        return jsonify({"reply": "Please type a message 🙂"})

    totals = fetch_user_totals(user)
    recent_tx = fetch_recent_transactions(user, limit=5)

    financial_context = f"""
Username: {user}
Total income: Rs. {totals['income_total']}
Income only: ₹{totals['income_only']}
Savings: ₹{totals['saving']}
Expense: ₹{totals['expense']}
Available balance after expenses and savings: Rs. {totals['balance']}
Recent transactions: {recent_tx}
"""

    reply = ask_gemini(msg, financial_context)

    return jsonify({"reply": reply})


@app.route('/transactions')
@login_required
def transactions():
    user = session['user']

    with get_conn(row_factory=sqlite3.Row) as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, description, amount, COALESCE(LOWER(type),'expense') AS type,
                   COALESCE(category,'Uncategorized') AS category,
                   COALESCE(subcategory,'') AS subcategory,
                   timestamp
            FROM expenses
            WHERE user = ?
            ORDER BY timestamp DESC
        """, (user,))
        rows = cur.fetchall()

    grouped = {}
    for row in rows:
        raw_timestamp = str(row["timestamp"] or "")
        try:
            parsed = datetime.fromisoformat(raw_timestamp.replace("Z", ""))
        except ValueError:
            parsed = datetime.now()
        month_key = parsed.strftime("%B %Y")
        grouped.setdefault(month_key, []).append({
            "id": row["id"],
            "description": row["description"] or "Transaction",
            "amount": round(float(row["amount"] or 0), 2),
            "type": row["type"],
            "category": row["category"],
            "subcategory": row["subcategory"],
            "stamp": parsed.strftime("%d %b %Y, %I:%M %p")
        })

    return render_template("transactions.html", username=user, grouped_transactions=grouped)


@app.route('/expense-galaxy')
@login_required
def expense_galaxy():
    user = session['user']
    totals = fetch_user_totals(user)
    category_colors = {item["name"]: item["color"] for item in EXPENSE_CATEGORIES}

    with get_conn(row_factory=sqlite3.Row) as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(category,'Uncategorized') AS category,
                   COALESCE(subcategory,'Other') AS subcategory,
                   COUNT(*) AS tx_count,
                   COALESCE(SUM(amount),0) AS total
            FROM expenses
            WHERE user = ? AND COALESCE(LOWER(type),'expense') = 'expense'
            GROUP BY COALESCE(category,'Uncategorized'), COALESCE(subcategory,'Other')
            ORDER BY total DESC
        """, (user,))
        rows = cur.fetchall()

    by_category = {}
    for row in rows:
        category = row["category"] or "Uncategorized"
        item = by_category.setdefault(category, {
            "category": category,
            "amount": 0.0,
            "count": 0,
            "color": category_colors.get(category, "#64748b"),
            "breakdown": []
        })
        amount = float(row["total"] or 0)
        count = int(row["tx_count"] or 0)
        item["amount"] += amount
        item["count"] += count
        item["breakdown"].append({
            "name": row["subcategory"] or "Other",
            "amount": round(amount, 2),
            "count": count
        })

    galaxy_data = sorted(by_category.values(), key=lambda item: item["amount"], reverse=True)
    for item in galaxy_data:
        item["amount"] = round(item["amount"], 2)

    return render_template(
        "expense_galaxy.html",
        username=user,
        balance=totals["balance"],
        galaxy_data=galaxy_data
    )


@app.route('/delete-entry/<int:entry_id>', methods=['POST'])
@login_required
def delete_entry(entry_id):
    user = session['user']

    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM expenses WHERE id = ? AND user = ?", (entry_id, user))
        conn.commit()

    flash("Entry deleted successfully.", "success")
    return redirect(url_for('dashboard'))
#------------upload_bank_file-----------
def extract_text_from_pdf(pdf_path):
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        print("PDF extraction error:", e)

    return text

ALLOWED_EXTENSIONS = {"pdf", "xls", "xlsx"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def import_transactions_from_excel(file_path, username):
    inserted_count = 0

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Header row is detected automatically
    header_row_index = None
    headers = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_values = [str(cell).strip().lower() if cell is not None else "" for cell in row]

        if any("date" in val for val in row_values) and (
            any("description" in val for val in row_values) or
            any("narration" in val for val in row_values) or
            any("transaction" in val for val in row_values)
        ):
            header_row_index = i
            headers = row_values
            break

    if header_row_index is None:
        return 0

    def find_col(possible_names):
        for idx, h in enumerate(headers):
            for name in possible_names:
                if name in h:
                    return idx
        return None

    date_col = find_col(["date"])
    desc_col = find_col(["transaction description", "description", "narration", "details", "reference"])
    credit_col = find_col(["credit", "deposit"])
    debit_col = find_col(["debit", "withdrawal"])

    if desc_col is None:
        return 0

    with get_conn() as conn:
        cursor = conn.cursor()

        for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not row:
                continue

            description = row[desc_col] if desc_col is not None and desc_col < len(row) else None
            raw_date = row[date_col] if date_col is not None and date_col < len(row) else None
            credit = row[credit_col] if credit_col is not None and credit_col < len(row) else None
            debit = row[debit_col] if debit_col is not None and debit_col < len(row) else None

            if not description:
                continue

            tx_type = None
            amount = None

            if credit not in [None, "", 0]:
                try:
                    amount = float(str(credit).replace(",", ""))
                    tx_type = "income"
                except:
                    pass
            elif debit not in [None, "", 0]:
                try:
                    amount = float(str(debit).replace(",", ""))
                    tx_type = "expense"
                except:
                    pass

            if amount is None or tx_type is None:
                continue

            if isinstance(raw_date, datetime):
                timestamp_value = raw_date.strftime("%Y-%m-%d")
            else:
                try:
                    parsed_date = pd.to_datetime(str(raw_date), errors="coerce")
                    if pd.isna(parsed_date):
                        timestamp_value = datetime.now().strftime("%Y-%m-%d")
                    else:
                        timestamp_value = parsed_date.strftime("%Y-%m-%d")
                except:
                    timestamp_value = datetime.now().strftime("%Y-%m-%d")

            cursor.execute("""
                INSERT INTO expenses (user, description, amount, category, emotion, type, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                username,
                str(description).strip(),
                amount,
                "Bank Import",
                "Neutral",
                tx_type,
                timestamp_value
            ))
            inserted_count += 1

        conn.commit()

    return inserted_count

def extract_text_from_pdf(pdf_path):
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        print("PDF extraction error:", e)

    return text

def parse_bank_statement_text(text):
    transactions = []
    lines = text.splitlines()

    for line in lines:
        line = line.strip()
        if not line:
            continue

        print("PDF LINE:", line)

        match = re.search(
            r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+(.+?)\s+(\d+(?:,\d{3})*(?:\.\d{1,2})?)\s*(DR|CR|Debit|Credit)?',
            line,
            re.IGNORECASE
        )

        if not match:
            continue

        date_str = match.group(1).strip()
        description = match.group(2).strip()

        try:
            amount = float(match.group(3).replace(",", ""))
        except:
            continue

        tx_flag = (match.group(4) or "").lower()

        if tx_flag in ["cr", "credit"]:
            tx_type = "income"
        else:
            tx_type = "expense"

        normalized_date = date_str.replace("/", "-")
        parts = normalized_date.split("-")

        try:
            if len(parts[2]) == 2:
                parts[2] = "20" + parts[2]
            day, month, year = parts
            formatted_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        except:
            formatted_date = datetime.now().strftime("%Y-%m-%d")

        transactions.append({
            "date": formatted_date,
            "description": description,
            "amount": amount,
            "type": tx_type
        })

    return transactions

def import_transactions_from_pdf(file_path, username):
    inserted_count = 0
    text = extract_text_from_pdf(file_path)

    print("===== PDF EXTRACTED TEXT =====")
    print(text)

    transactions = parse_bank_statement_text(text)

    print("===== PDF PARSED TRANSACTIONS =====")
    print(transactions)

    with get_conn() as conn:
        cursor = conn.cursor()

        for tx in transactions:
            timestamp_value = tx["date"] if tx["date"] else datetime.now().strftime("%Y-%m-%d")

            cursor.execute("""
                INSERT INTO expenses (user, description, amount, category, emotion, type, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                username,
                tx["description"],
                tx["amount"],
                "Bank Import",
                "Neutral",
                tx["type"],
                timestamp_value
            ))
            inserted_count += 1

        conn.commit()

    return inserted_count
    
@app.route('/upload-bank-file', methods=['POST'])
@login_required
def upload_bank_file():
    if 'bank_file' not in request.files:
        flash("No file selected.", "danger")
        return redirect(url_for('profile'))

    file = request.files['bank_file']

    if file.filename == '':
        flash("Please choose a file.", "warning")
        return redirect(url_for('profile'))

    if not allowed_file(file.filename):
        flash("Only PDF, XLS, and XLSX files are allowed.", "danger")
        return redirect(url_for('profile'))

    filename = secure_filename(file.filename)
    filename = f"{session['user']}_{filename}"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(save_path)

    ext = filename.rsplit('.', 1)[1].lower()
    inserted_count = 0

    print("===== FILE SAVED =====")
    print("Saved path:", save_path)

    try:
        if ext in ['xls', 'xlsx']:
            inserted_count = import_transactions_from_excel(save_path, session['user'])

            if inserted_count > 0:
                flash(f"Excel uploaded successfully! {inserted_count} transactions imported.", "success")
                return redirect(url_for('dashboard'))
            else:
                flash("Excel uploaded, but no valid transactions were found.", "warning")
                return redirect(url_for('profile'))

        elif ext == 'pdf':
            inserted_count = import_transactions_from_pdf(save_path, session['user'])

            if inserted_count > 0:
                flash(f"PDF uploaded successfully! {inserted_count} transactions imported.", "success")
                return redirect(url_for('dashboard'))
            else:
                flash("PDF uploaded, but no valid transactions were detected.", "warning")
                return redirect(url_for('profile'))

        else:
            flash("Unsupported file type.", "danger")
            return redirect(url_for('profile'))

    except Exception as e:
        print("IMPORT ERROR:", e)
        flash(f"Import failed: {str(e)}", "danger")
        return redirect(url_for('profile'))
    


def local_finance_advice(user_message, financial_context):
    message = user_message.lower()
    if any(word in message for word in ["hi", "hello", "hey"]):
        return "Hi! I am your FinGuard finance coach. Ask me about saving, reducing expenses, budgeting, or planning a goal."
    if "save" in message or "saving" in message:
        return (
            "Start with a realistic saving target from your available balance. "
            "A good plan is: essentials first, then fixed savings, then wants. "
            "If savings is more than your balance, reduce the goal or add more income first."
        )
    if "budget" in message or "plan" in message:
        return (
            "Try a simple split: 50% needs, 30% wants, 20% savings. "
            "If expenses are high, review Food, Transport, Shopping, and subscriptions first because they are easiest to adjust."
        )
    if "spend" in message or "expense" in message:
        return (
            "Look for repeat spending and small daily costs. "
            "Set weekly limits for food delivery, shopping, and cab apps, then compare every Sunday."
        )
    return (
        "I can help with finance, savings, spending awareness, budgets, and FinGuard data. "
        "Tell me your goal or ask where you can reduce expenses."
    )


def ask_gemini(user_message, financial_context):
    api_key = GEMINI_API_KEY
    if not api_key:
        return local_finance_advice(user_message, financial_context)

    try:
        genai.configure(api_key=api_key)

        prompt = f"""
You are FinGuard AI, a warm personal finance consultant inside a personal expense tracking app.

User financial data:
{financial_context}

User question:
{user_message}

Rules:
- Behave like a practical finance and savings coach, similar to a concise ChatGPT for money.
- Give personalized advice using the user's income, expenses, savings, balance, and recent transactions.
- Explain unsafe plans clearly, for example trying to save more money than the available balance.
- Suggest saving plans, spending awareness, category-level improvements, and next actions.
- Keep the answer clear, supportive, and practical.
- If the question is unrelated to the app, politely say you help with FinGuard and personal finance only.
"""

        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(prompt)

        return response.text if getattr(response, "text", "") else local_finance_advice(user_message, financial_context)

    except Exception as e:
        print("Gemini error:", e)
        return local_finance_advice(user_message, financial_context)

@app.route('/add-goal', methods=['POST'])
@login_required
def add_goal():
    user = session['user']
    goal_name = request.form['goal_name']
    target_amount = float(request.form['target_amount'])
    target_date = request.form['target_date']

    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO goals (user, goal_name, target_amount, target_date)
            VALUES (?, ?, ?, ?)
        """, (user, goal_name, target_amount, target_date))
        conn.commit()

    flash("Goal added successfully!", "success")
    return redirect(url_for('goal_planner'))

@app.route('/delete-goal/<int:goal_id>', methods=['POST'])
@login_required
def delete_goal(goal_id):
    user = session['user']

    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM goals
            WHERE id = ? AND user = ?
        """, (goal_id, user))
        conn.commit()

    flash("Goal deleted successfully!", "success")
    return redirect(url_for('goal_planner'))


@app.route('/goal-summary')
@login_required
def goal_quick():
    user = session['user']
    goals = build_goal_rows(user)
    return render_template("goal_summary.html", goals=goals, username=user)


@app.route('/tracker')
@login_required
def tracker():
    user = session['user']
    period = (request.args.get("period") or "week").lower()
    if period not in {"week", "month", "year"}:
        period = "week"

    today = datetime.today().date()
    selected_week = request.args.get("week") or "this"
    selected_month = request.args.get("month") or str(today.month)
    selected_year = request.args.get("year") or str(today.year)

    if period == "week":
        current_week = today.isocalendar().week
        try:
            week_number = int(selected_week)
        except (TypeError, ValueError):
            week_number = current_week - 1 if selected_week == "previous" else current_week
        week_number = max(1, min(52, week_number))
        selected_week = selected_week if selected_week in {"this", "previous"} else str(week_number)
        start_date = datetime.fromisocalendar(today.year, week_number, 1).date()
        end_date = start_date + timedelta(days=6)
        if selected_week == "this":
            label = "This week"
        elif selected_week == "previous":
            label = "Previous week"
        else:
            label = f"Week {week_number}"
    elif period == "month":
        try:
            month_number = int(selected_month)
        except (TypeError, ValueError):
            month_number = today.month
        month_number = max(1, min(12, month_number))
        selected_month = str(month_number)
        start_date = today.replace(month=month_number, day=1)
        if month_number == 12:
            end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = today.replace(month=month_number + 1, day=1) - timedelta(days=1)
        label = start_date.strftime("%B %Y")
    else:
        try:
            year_number = int(selected_year)
        except (TypeError, ValueError):
            year_number = today.year
        year_number = max(today.year - 10, min(today.year + 1, year_number))
        selected_year = str(year_number)
        start_date = datetime(year_number, 1, 1).date()
        end_date = datetime(year_number, 12, 31).date()
        label = str(year_number)

    with get_conn(row_factory=sqlite3.Row) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT description, amount,
                   COALESCE(category,'Uncategorized') AS category,
                   COALESCE(subcategory,'Other') AS subcategory,
                   timestamp
            FROM expenses
            WHERE user = ? AND COALESCE(LOWER(type),'expense') = 'expense'
            ORDER BY timestamp DESC
        """, (user,))
        rows = cursor.fetchall()

    tracker_map = {}
    total_amount = 0.0
    for row in rows:
        raw_timestamp = str(row["timestamp"] or "")
        try:
            parsed = datetime.fromisoformat(raw_timestamp.replace("Z", ""))
        except ValueError:
            continue

        if parsed.date() < start_date or parsed.date() > end_date:
            continue

        amount = float(row["amount"] or 0)
        total_amount += amount
        subcategory = row["subcategory"] or row["category"] or "Other"
        bucket = tracker_map.setdefault(subcategory, {
            "subcategory": subcategory,
            "category": row["category"] or "Uncategorized",
            "total": 0.0,
            "entries": []
        })
        bucket["total"] += amount
        bucket["entries"].append({
            "date": parsed.strftime("%d %b %Y"),
            "time": parsed.strftime("%I:%M %p"),
            "description": row["description"] or subcategory,
            "amount": round(amount, 2)
        })

    tracker_rows = sorted(tracker_map.values(), key=lambda item: item["total"], reverse=True)
    for item in tracker_rows:
        item["total"] = round(item["total"], 2)

    return render_template(
        "tracker.html",
        username=user,
        period=period,
        period_label=label,
        total_amount=round(total_amount, 2),
        tracker_rows=tracker_rows,
        selected_week=selected_week,
        selected_month=str(selected_month),
        selected_year=str(selected_year),
        week_options=[("this", "This week"), ("previous", "Previous week")] + [(str(number), f"Week {number}") for number in range(1, 53)],
        month_options=[(str(number), datetime(today.year, number, 1).strftime("%B")) for number in range(1, 13)],
        year_options=[str(year) for year in range(today.year - 5, today.year + 2)]
    )


@app.route('/goal-planner')
@login_required
def goal_planner():
    user = session['user']
    goals = build_goal_rows(user)

    return render_template("goal_planner.html", goals=goals, username=user)



if __name__ == '__main__':
    app.run(debug=False)
